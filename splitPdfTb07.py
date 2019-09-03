import PyPDF2
import re
import sys, os
import datetime
import win32com.client as win32
import openpyxl
import win32api
import logging
###### ver 3 ********
#test 23ph tach 941 file
#read/write data vao file temp\mau.xlsx. Neu danh ba email ko du mst thi se them vao danh sach

#ham kiem tra dinh dang file
#input: file_name, exam: file01.pdf
#output: file_type: .pdf
def file_type(str_file_name):
    return str_file_name[-4:]


#ham cat file pdf
#input: file_pdf_path, directory_name
#output: du_lieu = [['mst','file_result_path'], ['mst','file_result_path']]
def split_pdf_tb07(pdf_file_path, directory_name):
    pdf_file = open(pdf_file_path,'rb') #mo file pdf: pdf_file_path
    pdf_reader = PyPDF2.PdfFileReader(pdf_file)
    pdf_writer = PyPDF2.PdfFileWriter()
    tong_so_trang_pdf = pdf_reader.numPages
    #print(pdf_reader.numPages) #tổng số trang

    #create folder for result
    try:
        os.mkdir(directory_name)
    #    #print('da tao folder ', directory_name)
    except FileExistsError:
    #    #print('.') #bo qua
        pass
    
    tim_thay_mst = False
    mst = ''
    #is_qln_07qln = False
    file_created = 0
    du_lieu = []
    for page_number in range(tong_so_trang_pdf):
        page_obj = pdf_reader.getPage(page_number)
        txt = page_obj.extractText()
        result_txt_07qln = txt.find('07/QLN')
        if (result_txt_07qln != -1):
            #mau 07/QLN o ngay trang dau tien
            #is_qln_07qln = True
            #dung la mau 07/QLN
            #f = open('filetext.txt','w+',encoding="utf-8")
            #f.write(txt)
            result = re.search('(Mãs)\n(thu)\n(:)(.*)',txt)
            #giong_mst_truoc = False
            if result:
                #tim thay
                tim_thay_mst = True
                #if (mst == result.group(4)):
                #    giong_mst_truoc = True
                mst = result.group(4)
                
                #tao file moi
                pdf_writer = PyPDF2.PdfFileWriter() #tạo đối tượng mới để có đc file mới, nếu ko bị giữ nguyên addPage trước đó
                pdf_writer.addPage(pdf_reader.getPage(page_number))
                #output_filename = '{}_page_{}.pdf'.format(fname, page+1)
                output_filename = directory_name+'\\'+'{}.pdf'.format(mst)
                print('Tạo file '+directory_name+'\\'+mst+'.pdf')
                with open(output_filename, 'wb') as out:
                    pdf_writer.write(out)
                    file_created = file_created + 1
                    du_lieu.append([mst, output_filename])
            '''else:
                #print('Khong tim thay mst')
                #ghi vao file truoc do da tao
                print(mst)
                if tim_thay_mst:
                    pdf_writer.addPage(pdf_reader.getPage(page_number))
                    output_filename = directory_name+'\\'+'{}.pdf'.format(mst)
                    print('add them vao file '+mst+'.pdf')
                    with open(output_filename, 'wb') as out:
                        pdf_writer.write(out)
            '''
        else:
            #neu ko tim thay 07/QLN nghia la trang thu 2 tro di
            if tim_thay_mst:
                pdf_writer.addPage(pdf_reader.getPage(page_number))
                output_filename = directory_name+'\\'+'{}.pdf'.format(mst)
                print('Thêm trang mới vào file '+mst+'.pdf')
                with open(output_filename, 'wb') as out:
                    pdf_writer.write(out)
    pdf_file.close()
    #if (is_qln_07qln == False):
    #    print('Đây không phải là mẫu thông báo 07/QLN nên ứng dụng không hỗ trợ')
    #else:
    #    print('Số file đã tạo: '+str(file_created))
    return du_lieu #end def split_pdf_tb07



#main
print('SplitPdfTb07 phiên bản 3.0')

str_datetime = str(datetime.datetime.now())[0:19].replace(':','-') #string: nam-thang-ngay gio-phut-giay
#get file's path in all arguments
all_files_path = []
file_path = ''
#pdf_file_path = sys.argv[1]
for index_arg in range(len(sys.argv)):
    if index_arg != 0:
        #index=0: la file .py dang chay nen phai bo qua
        #print(sys.argv[index_arg])
        if (file_type(sys.argv[index_arg]) == '.pdf'):
            all_files_path.append(sys.argv[index_arg])
            pdf_file_path = sys.argv[index_arg]
            file_path_tmp = pdf_file_path.rfind('\\') #tim ky tu \ tu phai sang
            if (file_path != '' and file_path != pdf_file_path[0:file_path_tmp]):
                print("Các file thông báo phải cùng 1 thư mục. Chương trình sẽ dừng hoạt động")
                os.system("pause")
                sys.exit()
            file_path = pdf_file_path[0:file_path_tmp]

try:
    os.mkdir(file_path+'\\log') #tao folder ghi file log
except FileExistsError:
    pass
file_log = file_path+"\\log\\log_"+str_datetime+".txt"
logging.basicConfig(filename=file_log, level=logging.DEBUG)

if len(all_files_path) < 1:
    print("Ứng dụng hiện chỉ hỗ trợ file pdf thông báo nợ mẫu 07. Các file bạn đưa vào không có file nào đáp ứng được. Nhấn phím bất kỳ để thoát khỏi chương trình chạy.")
    os.system("pause")
    #os.system("exit")
    sys.exit()

#ktra co file mau.xlsx hay ko
is_exists_file_mau = os.path.isfile(file_path+'\\mau.xlsx')
if not is_exists_file_mau:
    print("Không có file mau.xlsx cùng thư mục với file thông báo. Chương trình sẽ dừng hoạt động")
    os.system("pause")
    sys.exit()

#ktra dinh dang file co phai pdf khong
#if (file_type(pdf_file_path) != '.pdf'):
#    print("Ứng dụng hiện chỉ hỗ trợ file pdf thông báo nợ mẫu 07")
#    os.system("pause")
#    #os.system("exit")
#    sys.exit()

#str_datetime = str(datetime.datetime.now())[0:19].replace(':','-') #string: nam-thang-ngay gio-phut-giay
#directory_name = file_path+'\\test' #str(datetime.datetime.now())[0:19].replace(':','-')
#create folder for result

du_lieu = [] #ghi toan bo ket qua tach file pdf ra day

#file_path = ''
for index_list in range(len(all_files_path)):
    pdf_file_path = all_files_path[index_list]
    print('Xử lý file '+pdf_file_path)
    logging.debug('Xu ly file '+pdf_file_path)
    file_path_tmp = pdf_file_path.rfind('\\') #tim ky tu \ tu phai sang
    file_path = pdf_file_path[0:file_path_tmp]
    #print(file_path)

    #str_datetime_excel = str(datetime.datetime.now())[0:19].replace(':','-') #string: nam-thang-ngay gio-phut-giay
    directory_name = file_path+'\\'+str_datetime #str(datetime.datetime.now())[0:19].replace(':','-')

    du_lieu = du_lieu + split_pdf_tb07(pdf_file_path, directory_name)

if (len(du_lieu) > 0):
    path_file_mau = file_path+'\\'+"Mau.xlsx"
    
    wb = openpyxl.load_workbook(filename = path_file_mau) #load file
    all_sheet_name = wb.sheetnames
    logging.debug('Mo file '+path_file_mau+' de lam viec')


	#ktra file mau phai co sheet "danhba"
    if ('danhba' not in all_sheet_name):
	    print("File mau.xlsx không có sheet danhba. Chương trình không thể hoạt động tiếp")
        #logging.debug('File mau.xls khong co sheet danhba. Chuong trinh khong hoat dong tiep')
	    os.system("pause")
		#os.system("exit")
		#sys.exit("File mau.xls không có sheet danhba. Chương trình không thể hoạt động tiếp")
	    sys.exit()
	
	#(3) move mau.xls to folder
    #shutil.move(file_mau, file_path+'\\temp\\Mau'+str_datetime+'.xls')
    #logging.debug('Move file mau.xls sang '+file_path+'\\temp\\Mau'+str_datetime+'.xls')

    sheet_name = wb['danhba']
    i = 1 #xác định số dòng ở file mau.xlsx
    del_mst = []

    logging.debug('Xoa du lieu file path o file Mau.xlsx')
    max_row = sheet_name.max_row #số dòng trên sheet danhba
    for row in sheet_name.iter_rows(min_row=2, min_col=6, max_row=max_row, max_col=6):
        for cell in row:
            cell.value = '' #cho cột F = ''

    #for row in sheet_name.values:
    #    if i > 1: #dòng tiêu đề của file excel nên bỏ qua
    #        sheet_name['F'+str(i)] = ''

    print('Đang ghi dữ liệu vào file danh bạ.')
    logging.debug('Ghi du lieu vao file Mau.xlsx')
    for row in sheet_name.values:
        for i1 in range(len(du_lieu)):
            if (str(row[0]) == str(du_lieu[i1][0])):
                sheet_name['F'+str(i)] = ''
                sheet_name['F'+str(i)] = du_lieu[i1][1]
                #logging.debug('----- Ghi vao F'+str(i)+': '+du_lieu[i1][1])
                del_mst.append(i1) #lưu index của list du_lieu
        i = i + 1
    logging.debug('Da ghi du lieu vao file Mau.xlsx')

    for i in range(len(du_lieu)):
        max_row = sheet_name.max_row+1
        if i not in del_mst:
            print('MST '+str(du_lieu[i][0])+' không nằm trong danh bạ của bạn. Chương trình sẽ ghi thêm dữ liệu mst này ở cuối danh sách, bạn phải thiết lập thêm các dữ liệu sau đó (nếu cần)')
            sheet_name['A'+str(max_row)] = du_lieu[i][0]
            sheet_name['F'+str(max_row)] = du_lieu[i][1]
    wb.save(path_file_mau)
    wb.close()
    logging.debug('Da ghi du lieu MST bi thieu vao file Mau.xlsx')

    #ghi duong dan file mau.xls vao 1 file txt ==> access se doc file nay
    #f = open('config.txt','w+',encoding="utf-8") #w+: mo de doc va ghi. ghi de neu ton tai, neu chua co thi tao moi
    #f.write(file_mau)
    #f.close()
    
    logging.debug('Tong so file da tao: '+str(len(du_lieu)))
    print('Tổng số file đã tạo: '+str(len(du_lieu)))
    print('Nhấn phím bất kỳ để thoát.')
    os.system("pause")
    #win32api.ShellExecute(0,'open',file_mau,'','',1) #open mau.xls
    sys.exit()
'''
for page_number in range(5):
    page_obj = pdf_reader.getPage(page_number)
    txt = page_obj.extractText()
    result = txt.find('07/QLN')
    if (result != -1):
        #print('tim thay')

    else:
        #print('khong tim thay')
'''