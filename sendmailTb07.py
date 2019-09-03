import win32com.client as win32
import sys, os
import openpyxl
import xlwings as xw

#ham kiem tra dinh dang file
#input: file_name, exam: file01.pdf
#output: file_type: .pdf
def file_type(str_file_name):
    return str_file_name[-4:]

print('SendemailTb07 phiên bản 3.0')
print('Chương trình chỉ gửi email cho những dòng có đầy đủ thông tin email và file đính kèm')
print('---------------------------')
print('Nhấn nút bất kỳ để xác nhận bạn muốn gửi. Nếu không hãy nhấn nút đóng chương trình lại...')
os.system("pause")

excel_file_path = sys.argv[1] #lay duong dan file truyen vao
#check file type
if (file_type(excel_file_path) != 'xlsx'):
    print("File của bạn không phải là file excel (.xlsx). Chương trình sẽ dừng hoạt động")
    os.system("pause")
    sys.exit()

#doc du lieu file excel
wb = openpyxl.load_workbook(filename = excel_file_path) #load file
all_sheet_name = wb.sheetnames
if ('danhba' not in all_sheet_name):
	print("File excel của bạn không có sheet danhba. Chương trình không thể hoạt động tiếp")
	os.system("pause")
	sys.exit()
sheet_name = wb['danhba']
max_row = sheet_name.max_row
wb.close

wb = xw.Book(excel_file_path) #open file excel by slwings
#print(wb.sheets['danhba'].range('D2').value)
sheet_name = wb.sheets['danhba'] #lam viec tren sheet danh ba
list_to_send = []

for i in range(2, max_row + 1):
    if (str(sheet_name.range('E'+str(i)).value).strip() != '' and str(sheet_name.range('F'+str(i)).value).strip() != '' and str(sheet_name.range('E'+str(i)).value).strip() != 'None' and str(sheet_name.range('F'+str(i)).value).strip() != 'None'):
        list_to_send.append({'email': sheet_name.range('E'+str(i)).value, 'subject': sheet_name.range('C'+str(i)).value, 'body': sheet_name.range('D'+str(i)).value, 'file': sheet_name.range('F'+str(i)).value}) #add column file_path to list
    #print(sheet_name.range('F'+str(i)).value)
#print(list_to_send)
#gui email
outlook = win32.Dispatch('outlook.application')
try:
    for item in list_to_send:
        mail = outlook.CreateItem(0)
        mail.To = item['email']
        #'lhdieu.qni@gdt.gov.vn'
        mail.Subject = item['subject']
        #mail.Body = 'Message body'
        mail.HTMLBody = item['body'] #this field is optional
        # To attach a file to the email (optional):
        attachment  = item['file']
        mail.Attachments.Add(attachment)
        mail.Send()
    print("Đã tạo và gửi yêu cầu gửi email cho outlook thành công")
    os.system("pause")
    sys.exit()
except Exception as e:
    print("Có lỗi khi tạo và gửi email. Bạn hãy kiểm tra lại có email nào không đúng định dạng không")
    os.system("pause")
    sys.exit()